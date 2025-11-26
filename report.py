
## Requiere tener instalado el módulo openpyxl
## pip install openpyxl


import pandas as pd
from openpyxl import load_workbook



def generate_report(vacias, casi_vacias):
    
    columnas = ["ID","NOMBRE"]
    fr1 = pd.DataFrame(vacias, columns=columnas)


    columnas = ["ID","NOMBRE","Nº","LISTA PARTICIPANTES","ACCESO MÁS RECIENTE"]
    fr2 = pd.DataFrame(casi_vacias, columns=columnas)

    with pd.ExcelWriter("salida.xlsx", engine="openpyxl") as writer:
        fr1.to_excel(writer, sheet_name="Vacías", index=False)
        fr2.to_excel(writer, sheet_name="Abandonadas", index=False)

    ajustar_columnas_excel("salida.xlsx")



def ajustar_columnas_excel(ruta_excel):
    """
    Ajusta el ancho de todas las columnas en todas las hojas de un archivo Excel
    en función del contenido más ancho encontrado en cada columna.
    """

    wb = load_workbook(ruta_excel)

    for hoja in wb.sheetnames:
        ws = wb[hoja]

        for columna in ws.columns:
            max_len = 0
            col_letter = columna[0].column_letter

            for celda in columna:
                try:
                    if celda.value:
                        max_len = max(max_len, len(str(celda.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_len + 1  # margen extra

    wb.save(ruta_excel)



def export_courses(vacias, casi_vacias):
    
    with open("aulas_vacias.csv", "w", encoding='utf-8') as vaciasfile:
        vaciasfile.write("ID;NOMBRE;URL\n")
        for a in vacias:
            vaciasfile.write(";".join(str(i) for i in a)+"\n")
        
    vaciasfile.close()

    with open("aulas_casi_vacias.csv", "w", encoding='utf-8') as casivaciasfile:
        casivaciasfile.write("ID;NOMBRE;URL;Nº PARTICIPANTES;LISTA PARTICIPANTES;ACCESO MÁS RECIENTE;\n")
        for a in casi_vacias:
            casivaciasfile.write(";".join(str(i) for i in a)+"\n")
        
    casivaciasfile.close()
        

